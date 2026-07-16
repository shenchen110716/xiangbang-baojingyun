"""Employment fact and import APIs (v4.2 §14.2).

Every mutation goes through the services: the router does no ORM writes and
re-derives no role logic, so employer scope and versioning cannot drift from
what Phase 1 and the fact service enforce.
"""
import io

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..core.audit import audit
from ..core.business_time import business_now
from ..core.db import db
from ..core.rbac import require_role
from ..core.security import current_user
from ..models import EmploymentFact, EmploymentFactMatch, EmploymentFeedbackBatch, InsuredPerson, User, WorkPosition
from ..schemas.employment import (
    BatchOut,
    FactCorrectIn,
    FactListOut,
    FactOut,
    ImportConfirmIn,
    ImportConfirmOut,
    ImportPreviewOut,
    ManualMatchIn,
)
from ..services.employer_scopes import allowed_employer_ids, assert_employer_access
from ..services.employment_facts import active_facts, correct_fact, serialize_fact
from ..services.employment_import import TEMPLATE_HEADER, confirm_import, preview_import

router = APIRouter(prefix="/api", tags=["employment-facts"])

# 用工事实属于企业用工数据，业务员无关；平台管理员可跨企业查看。
_ENTERPRISE_OR_ADMIN = require_role("admin", "enterprise", detail="无权访问用工事实数据")


def _enterprise_id(user: User, requested: int | None = None) -> int:
    if user.role == "enterprise":
        if not user.enterprise_id:
            raise HTTPException(403, "账号未绑定投保单位")
        return user.enterprise_id
    if not requested:
        raise HTTPException(400, "平台管理员必须指定投保单位")
    return requested


@router.get("/employment-feedback/template", dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def employment_template(user: User = Depends(current_user)):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "真实用工反馈模板"
    sheet.append(TEMPLATE_HEADER)
    sheet.append(["项目 A", "E001", "张三", "340123199001011238",
                  "2026-03-01", "", "2026-03-02", "EXT-1", ""])
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DCE6FF")
    for column, width in {"A": 24, "B": 16, "C": 14, "D": 23, "E": 20,
                          "F": 20, "G": 20, "H": 20, "I": 24}.items():
        sheet.column_dimensions[column].width = width
    for row in range(2, sheet.max_row + 1):
        # Text format, or Excel turns the ID into 3.40123E+17.
        sheet.cell(row, 4).number_format = "@"
    output = io.BytesIO()
    book.save(output)
    book.close()
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employment-feedback-template.xlsx"},
    )


@router.post("/employment-feedback/import/preview", response_model=ImportPreviewOut,
             dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
async def import_preview(file: UploadFile = File(...),
                         enterprise_id: int | None = Query(None),
                         user: User = Depends(current_user),
                         session: Session = Depends(db)):
    content = await file.read()
    result = preview_import(session, user,
                            enterprise_id=_enterprise_id(user, enterprise_id),
                            filename=file.filename or "", content=content)
    session.commit()
    return result


@router.post("/employment-feedback/import/confirm", response_model=ImportConfirmOut,
             dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def import_confirm(data: ImportConfirmIn, user: User = Depends(current_user),
                   session: Session = Depends(db)):
    result = confirm_import(session, user, batch_id=data.batch_id,
                            confirm_token=data.confirm_token)
    session.commit()
    audit(session, user, "import_confirm", "employment_batch", str(data.batch_id),
          f"created_facts={result['created_facts']}")
    return result


def _batch_scope(session: Session, user: User):
    stmt = select(EmploymentFeedbackBatch)
    if user.role == "enterprise":
        stmt = stmt.where(EmploymentFeedbackBatch.enterprise_id == user.enterprise_id)
    return stmt


@router.get("/employment-feedback/batches", response_model=list[BatchOut],
            dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def list_batches(user: User = Depends(current_user), session: Session = Depends(db)):
    stmt = _batch_scope(session, user).order_by(EmploymentFeedbackBatch.id.desc())
    return list(session.scalars(stmt))


@router.get("/employment-feedback/batches/{item_id}", response_model=BatchOut,
            dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def get_batch(item_id: int, user: User = Depends(current_user),
              session: Session = Depends(db)):
    batch = session.get(EmploymentFeedbackBatch, item_id)
    if not batch:
        raise HTTPException(404, "导入批次不存在")
    if user.role == "enterprise" and batch.enterprise_id != user.enterprise_id:
        raise HTTPException(403, "无权查看该批次")
    return batch


@router.get("/employment-facts", response_model=FactListOut,
            dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def list_facts(user: User = Depends(current_user), session: Session = Depends(db)):
    return {"items": [serialize_fact(f) for f in active_facts(session, user)]}


@router.get("/employment-facts/unmatched", response_model=FactListOut,
            dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def list_unmatched(user: User = Depends(current_user), session: Session = Depends(db)):
    """Facts that are real but not yet bound to a person, so they stay out of
    published metrics until a human resolves them (§20.6)."""
    stmt = select(EmploymentFact).where(EmploymentFact.status == "pending_match")
    allowed = allowed_employer_ids(session, user)
    if allowed is not None:
        if not allowed:
            return {"items": []}
        stmt = stmt.where(EmploymentFact.actual_employer_id.in_(allowed))
    if user.role == "enterprise":
        stmt = stmt.where(EmploymentFact.enterprise_id == user.enterprise_id)
    return {"items": [serialize_fact(f) for f in session.scalars(stmt.order_by(EmploymentFact.id))]}


@router.get("/employment-facts/{item_id}", response_model=FactOut,
            dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def get_fact(item_id: int, user: User = Depends(current_user),
             session: Session = Depends(db)):
    fact = session.get(EmploymentFact, item_id)
    if not fact:
        raise HTTPException(404, "用工事实不存在")
    if user.role == "enterprise" and fact.enterprise_id != user.enterprise_id:
        raise HTTPException(403, "无权查看该用工事实")
    assert_employer_access(session, user, fact.actual_employer_id)
    return serialize_fact(fact)


@router.patch("/employment-facts/{item_id}/correct", response_model=FactOut,
              dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def correct(item_id: int, data: FactCorrectIn, user: User = Depends(current_user),
            session: Session = Depends(db)):
    fact = correct_fact(session, user, item_id,
                        actual_hire_at=data.actual_hire_at,
                        actual_leave_at=data.actual_leave_at,
                        reason=data.reason)
    payload = serialize_fact(fact)
    session.commit()
    audit(session, user, "correct", "employment_fact", str(item_id),
          f"new_revision={payload['revision_no']} reason={data.reason}")
    return payload


@router.post("/employment-facts/unmatched/{item_id}/match", response_model=FactOut,
             dependencies=[Depends(_ENTERPRISE_OR_ADMIN)])
def manual_match(item_id: int, data: ManualMatchIn, user: User = Depends(current_user),
                 session: Session = Depends(db)):
    fact = session.get(EmploymentFact, item_id)
    if not fact:
        raise HTTPException(404, "用工事实不存在")
    if fact.status != "pending_match":
        raise HTTPException(409, "只能手工匹配待匹配的用工事实")
    if user.role == "enterprise" and fact.enterprise_id != user.enterprise_id:
        raise HTTPException(403, "无权操作该用工事实")
    assert_employer_access(session, user, fact.actual_employer_id)

    person = session.get(InsuredPerson, data.person_id)
    if not person or person.enterprise_id != fact.enterprise_id:
        raise HTTPException(400, "被保险人不属于该投保单位")
    position = session.get(WorkPosition, person.position_id) if person.position_id else None
    # Binding a person from another project would move the fact out of the
    # operator's own scope, so hold the same employer line as the auto ladder.
    if not position or position.actual_employer_id != fact.actual_employer_id:
        raise HTTPException(400, "被保险人不属于该实际工作单位")

    fact.person_id = person.id
    fact.status = "active"
    session.add(EmploymentFactMatch(
        employment_fact_id=fact.id, match_status="matched", match_method="manual",
        matched_person_id=person.id, candidate_person_id=person.id, confidence=1.0,
        reason=(data.reason or "人工匹配")[:255], confirmed_by=user.id,
        confirmed_at=business_now(), created_at=business_now()))
    payload = serialize_fact(fact)
    session.commit()
    audit(session, user, "manual_match", "employment_fact", str(item_id),
          f"person_id={person.id}")
    return payload
