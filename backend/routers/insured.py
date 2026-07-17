import csv
import io
from datetime import datetime
from typing import Literal

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from ..core.audit import audit
from ..core.business_time import as_business_time
from ..core.db import db
from ..core.id_number import is_valid_id_number
from ..core.security import current_user
from ..models import ActualEmployer, AgentCommission, Enterprise, InsurancePlan, InsuredPerson, Policy, PolicyMember, User, WorkPosition
from ..schemas import BulkPersonIn, PersonIn, PersonUpdate
from ..services import activate_person_policy, allowed_employer_ids, assert_employer_access, correct_person_policy_dates, effective_person_status, is_enterprise_owner, plan_price_for_class, pricing_snapshot, require_usage_funded, serialize, strip_internal_pricing, terminate_person_policy
from ..services.spreadsheet import MAX_IMPORT_FILE_BYTES, read_import_rows
from ..services.timeliness_recalc import record_operation

router = APIRouter(prefix="/api", tags=["insured"])


def _parse_business_time(raw: str | None, label: str) -> datetime | None:
    if raw is None or not raw.strip():
        return None
    try:
        value = datetime.fromisoformat(raw.strip().replace("Z", "+00:00"))
    except ValueError as exc:
        raise HTTPException(400, f"{label}时间格式不正确，应为 yyyy-MM-dd") from exc
    return as_business_time(value)


def _coverage_dates(session: Session, person_id: int) -> tuple[datetime | None, datetime | None]:
    member = session.scalar(select(PolicyMember).where(PolicyMember.person_id == person_id).order_by(PolicyMember.id.desc()))
    return (member.effective_at, member.terminated_at) if member else (None, None)


def _person_payload(session: Session, item: InsuredPerson) -> dict:
    payload = serialize(item)
    effective_at, terminated_at = _coverage_dates(session, item.id)
    payload["effective_at"], payload["terminated_at"] = effective_at, terminated_at
    payload["status"] = effective_person_status(item, terminated_at)
    return payload


def _person_employer_access(
    session: Session, user: User, person: InsuredPerson
) -> WorkPosition | None:
    position = session.get(WorkPosition, person.position_id) if person.position_id else None
    if position and position.actual_employer_id is not None:
        assert_employer_access(session, user, position.actual_employer_id)
        return position
    if user.role == "enterprise":
        if person.enterprise_id != user.enterprise_id:
            raise HTTPException(403, "无权访问其他企业员工")
        if not is_enterprise_owner(user):
            raise HTTPException(403, "员工未关联实际工作单位，项目负责人无权访问")
    elif user.role != "admin":
        raise HTTPException(403, "无权访问参保员工")
    return position


@router.get("/insured")
def insured(q: str = "", user: User = Depends(current_user), session: Session = Depends(db)):
    stmt = select(InsuredPerson).order_by(InsuredPerson.id.desc())
    if user.role=='enterprise' and user.enterprise_id:
        stmt=stmt.where(InsuredPerson.enterprise_id==user.enterprise_id)
        allowed=allowed_employer_ids(session,user)
        if allowed is not None:
            stmt=stmt.join(WorkPosition,InsuredPerson.position_id==WorkPosition.id).where(WorkPosition.actual_employer_id.in_(allowed))
    elif user.role!='admin': raise HTTPException(403,'无权查看参保员工')
    if q: stmt = stmt.where(or_(InsuredPerson.name.contains(q), InsuredPerson.phone.contains(q)))
    result=[]
    for x in session.scalars(stmt):
        item=_person_payload(session,x)
        # x.policy_id is cleared the moment a stop is scheduled (even a
        # future-dated 临时日结 auto-expiry) — once effective_person_status
        # has decided the person is still actually active, fall back to the
        # still-open PolicyMember's policy_id so 保险产品/保单号 don't go
        # blank while the row still says 在保.
        policy_id=x.policy_id
        if policy_id is None and item['status']=='active':
            latest_member=session.scalar(select(PolicyMember).where(PolicyMember.person_id==x.id).order_by(PolicyMember.id.desc()))
            policy_id=latest_member.policy_id if latest_member else None
        enterprise=session.get(Enterprise,x.enterprise_id);position=session.get(WorkPosition,x.position_id) if x.position_id else None;employer=session.get(ActualEmployer,position.actual_employer_id) if position and position.actual_employer_id else None;plan=session.get(InsurancePlan,position.plan_id) if position and position.plan_id else None;policy=session.get(Policy,policy_id) if policy_id else None
        relation=session.scalar(select(AgentCommission).where(AgentCommission.enterprise_id==x.enterprise_id,AgentCommission.plan_id==plan.id,AgentCommission.status=='active').order_by(AgentCommission.id.desc())) if plan else None
        item.update(enterprise_name=enterprise.name if enterprise else '',position_name=position.name if position else x.occupation,actual_employer_name=employer.name if employer else (position.actual_employer if position else ''),plan_id=plan.id if plan else None,plan_name=plan.name if plan else '',insurer=plan.insurer if plan else '',policy_no=policy.policy_no if policy else '',policy_status=policy.status if policy else '',effective_mode=plan.effective_mode if plan else '',billing_mode=plan.billing_mode if plan else '',**(pricing_snapshot(plan,relation,plan_price_for_class(session,plan,x.occupation_class)) if plan else {}))
        result.append(strip_internal_pricing(item,user))
    return result

@router.post("/insured")
def add_person(data: PersonIn, user: User = Depends(current_user), session: Session = Depends(db)):
    enterprise = session.get(Enterprise, data.enterprise_id)
    if not enterprise: raise HTTPException(404, "企业不存在")
    if user.role=="enterprise" and user.enterprise_id!=data.enterprise_id: raise HTTPException(403,"无权操作该单位")
    require_usage_funded(session, enterprise, user)
    if not is_valid_id_number(data.id_number): raise HTTPException(400,'身份证号格式不正确')
    if session.scalar(select(InsuredPerson.id).where(InsuredPerson.enterprise_id==data.enterprise_id,InsuredPerson.id_number==data.id_number).limit(1)): raise HTTPException(409,'该身份证号已在本单位参保，请勿重复添加')
    effective_at = _parse_business_time(data.effective_at, "生效")
    terminated_at = _parse_business_time(data.terminated_at, "停保")
    payload=data.model_dump(exclude={"effective_at", "terminated_at"})
    if data.position_id:
        position=session.get(WorkPosition,data.position_id)
        if not position or position.enterprise_id!=data.enterprise_id or position.status!='approved': raise HTTPException(400,"只能选择本单位已审核通过的有效岗位")
        if position.actual_employer_id is None:
            if user.role=='enterprise' and not is_enterprise_owner(user): raise HTTPException(403,'岗位未关联实际工作单位，项目负责人无权操作')
        else: assert_employer_access(session,user,position.actual_employer_id)
        payload['occupation']=position.name; payload['occupation_class']=position.occupation_class
    elif user.role=='enterprise' and not is_enterprise_owner(user):
        raise HTTPException(403,'项目负责人新增员工必须选择授权范围内的岗位')
    item = InsuredPerson(**payload); session.add(item); session.flush()
    member = correct_person_policy_dates(session, item, effective_at, terminated_at)
    if member is not None: item.status = "stopped" if member.terminated_at is not None else "active"
    record_operation(session, user=user, person=item, operation_type="enrollment")
    if member is not None and member.terminated_at is not None:
        record_operation(session, user=user, person=item, operation_type="termination")
    session.commit(); session.refresh(item); audit(session, user, "create", "insured_person", str(item.id)); return _person_payload(session, item)

@router.patch("/insured/{item_id}")
def update_person(item_id:int,data:PersonUpdate,user:User=Depends(current_user),session:Session=Depends(db)):
    item=session.get(InsuredPerson,item_id)
    if not item: raise HTTPException(404,'参保员工不存在')
    _person_employer_access(session,user,item)
    require_usage_funded(session, session.get(Enterprise, item.enterprise_id), user)
    values=data.model_dump(exclude_unset=True)
    if 'id_number' in values and values['id_number']!=item.id_number:
        if not is_valid_id_number(values['id_number']): raise HTTPException(400,'身份证号格式不正确')
        if session.scalar(select(InsuredPerson.id).where(InsuredPerson.enterprise_id==item.enterprise_id,InsuredPerson.id_number==values['id_number'],InsuredPerson.id!=item.id).limit(1)): raise HTTPException(409,'该身份证号已在本单位参保，请勿重复添加')
    if 'position_id' in values:
        position=session.get(WorkPosition,values['position_id'])
        if not position or position.enterprise_id!=item.enterprise_id or position.status!='approved': raise HTTPException(400,'只能选择本单位已审核通过的有效岗位')
        if position.actual_employer_id is None:
            if user.role=='enterprise' and not is_enterprise_owner(user): raise HTTPException(403,'岗位未关联实际工作单位，项目负责人无权操作')
        else: assert_employer_access(session,user,position.actual_employer_id)
        item.position_id=position.id;item.occupation=position.name;item.occupation_class=position.occupation_class
    for key in ('name','phone','id_number'):
        if key in values and values[key] is not None: setattr(item,key,values[key])
    effective_at = _parse_business_time(values.get('effective_at'), '生效') if 'effective_at' in values else None
    terminated_at = _parse_business_time(values.get('terminated_at'), '停保') if 'terminated_at' in values else None
    if effective_at is not None or terminated_at is not None:
        member = correct_person_policy_dates(session, item, effective_at, terminated_at)
        if member is not None: item.status = 'stopped' if member.terminated_at is not None else 'active'
        record_operation(session, user=user, person=item,
                         operation_type='termination' if terminated_at is not None else 'enrollment')
    session.commit();audit(session,user,'update','insured_person',str(item.id));return _person_payload(session,item)

@router.patch("/insured/{item_id}/status")
def insured_status(item_id:int,status_value:Literal["active","stopped","pending"]=Query(...,alias="status"),user:User=Depends(current_user),session:Session=Depends(db)):
    item=session.get(InsuredPerson,item_id)
    if not item: raise HTTPException(404,"参保员工不存在")
    _person_employer_access(session,user,item)
    require_usage_funded(session, session.get(Enterprise, item.enterprise_id), user)
    previous_status=item.status
    if status_value=="active" and previous_status!="active":
        activate_person_policy(session,item)
        record_operation(session, user=user, person=item, operation_type="enrollment")
    elif previous_status=="active" and status_value!="active":
        terminate_person_policy(session,item)
        record_operation(session, user=user, person=item, operation_type="termination")
    item.status=status_value
    session.commit();audit(session,user,"status_change","insured_person",str(item.id),status_value);return _person_payload(session,item)

@router.get("/insured/{item_id}/policy-members")
def insured_policy_members(item_id: int, user: User = Depends(current_user), session: Session = Depends(db)):
    item = session.get(InsuredPerson, item_id)
    if not item: raise HTTPException(404, "参保员工不存在")
    _person_employer_access(session,user,item)
    rows = session.scalars(select(PolicyMember).where(PolicyMember.person_id == item_id).order_by(PolicyMember.id.desc()))
    result = []
    for pm in rows:
        policy = session.get(Policy, pm.policy_id)
        plan = session.get(InsurancePlan, policy.plan_id) if policy else None
        entry = serialize(pm)
        entry.update(policy_no=policy.policy_no if policy else "", insurer=plan.insurer if plan else "", plan_name=plan.name if plan else "", effective_mode=plan.effective_mode if plan else "")
        result.append(entry)
    return result

@router.get("/insured/import-template")
def insured_import_template(user:User=Depends(current_user)):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = '批量导入模板'
    sheet.append(['姓名', '身份证号', '手机号', '投保单位', '实际工作单位', '岗位名称', '生效日期', '停保日期'])
    sheet.append(['张三', '340123199001011234', '13800000000', '', '', '', '2026-01-01', ''])
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='DCE6FF')
    for column, width in {'A': 14, 'B': 23, 'C': 16, 'D': 24, 'E': 24, 'F': 20, 'G': 16, 'H': 16}.items():
        sheet.column_dimensions[column].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).number_format = '@'
        sheet.cell(row, 3).number_format = '@'
    output = io.BytesIO()
    book.save(output)
    book.close()
    filename = 'insured-import-template.xlsx'
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return StreamingResponse(iter([output.getvalue()]), media_type=media_type, headers={'Content-Disposition': f'attachment; filename={filename}'})


def _read_import_rows(content: bytes, filename: str) -> list[list[str]]:
    # Shared with the employment-fact import; see services/spreadsheet.py.
    return read_import_rows(content, filename)

@router.post("/insured/bulk")
def bulk_add_people(data:BulkPersonIn,user:User=Depends(current_user),session:Session=Depends(db)):
    if user.role=='enterprise' and user.enterprise_id!=data.enterprise_id: raise HTTPException(403,'无权操作该单位')
    enterprise=session.get(Enterprise,data.enterprise_id)
    if not enterprise: raise HTTPException(404,'投保单位不存在')
    require_usage_funded(session, enterprise, user)
    position=session.get(WorkPosition,data.position_id)
    if not position or position.enterprise_id!=data.enterprise_id or position.status!='approved': raise HTTPException(400,'只能选择本单位已审核通过的岗位')
    if position.actual_employer_id is None:
        if user.role=='enterprise' and not is_enterprise_owner(user): raise HTTPException(403,'岗位未关联实际工作单位，项目负责人无权操作')
    else: assert_employer_access(session,user,position.actual_employer_id)
    errors=[];created=[];seen=set()
    for index,row in enumerate(data.rows,start=2):
        identity=row.id_number.strip();name=row.name.strip()
        if not name or not identity: errors.append({'row':index,'message':'姓名和身份证号必填'});continue
        if identity in seen or session.scalar(select(InsuredPerson.id).where(InsuredPerson.id_number==identity).limit(1)): errors.append({'row':index,'message':'身份证号重复'});continue
        seen.add(identity);item=InsuredPerson(enterprise_id=data.enterprise_id,position_id=position.id,name=name,id_number=identity,phone=row.phone.strip(),occupation=position.name,occupation_class=position.occupation_class,status='pending');session.add(item);created.append(item)
    if errors: session.rollback();return {'ok':False,'created':0,'errors':errors}
    session.flush()
    for item in created:
        record_operation(session, user=user, person=item, operation_type='enrollment')
    session.commit()
    for item in created: session.refresh(item)
    audit(session,user,'bulk_create','insured_person',','.join(str(x.id) for x in created),f'count={len(created)}')
    return {'ok':True,'created':len(created),'errors':[],'ids':[x.id for x in created]}

@router.post("/insured/import-file")
async def import_insured_file(kind:Literal['enrollment','termination']=Form(...),enterprise_id:int=Form(...),position_id:int=Form(0),file:UploadFile=File(...),user:User=Depends(current_user),session:Session=Depends(db)):
    if user.role=='enterprise' and user.enterprise_id!=enterprise_id: raise HTTPException(403,'无权操作该单位')
    primary_enterprise=session.get(Enterprise,enterprise_id)
    if not primary_enterprise: raise HTTPException(404,'投保单位不存在')
    require_usage_funded(session, primary_enterprise, user)
    default_position=None
    if kind=='enrollment' and position_id:
        default_position=session.get(WorkPosition,position_id)
        if not default_position or default_position.enterprise_id!=enterprise_id or default_position.status!='approved': raise HTTPException(400,'批量参保必须选择本单位已审核通过的岗位')
        if default_position.actual_employer_id is None:
            if user.role=='enterprise' and not is_enterprise_owner(user): raise HTTPException(403,'岗位未关联实际工作单位，项目负责人无权操作')
        else: assert_employer_access(session,user,default_position.actual_employer_id)
    if file.size is not None and file.size > MAX_IMPORT_FILE_BYTES: raise HTTPException(413,'单个导入文件不能超过 10MB，请拆分后重试')
    content=await file.read();raw=_read_import_rows(content,file.filename or '')
    if len(raw)<2: raise HTTPException(400,'电子表格没有可导入的数据')
    headers={x.replace(' ',''):i for i,x in enumerate(raw[0])}
    name_col=headers.get('姓名');id_col=headers.get('身份证号');phone_col=headers.get('手机号')
    # 三列可选：留空则沿用上传时选择的默认投保单位/岗位（兼容旧模板和小程序端），
    # 填写则按名称匹配，用于单次导入多个不同单位/岗位的名单（见反馈条目 5）。
    enterprise_col=headers.get('投保单位');employer_col=headers.get('实际工作单位');position_col=headers.get('岗位名称')
    effective_col=headers.get('生效日期');terminated_col=headers.get('停保日期')
    if id_col is None or (kind=='enrollment' and name_col is None): raise HTTPException(400,'模板必须包含姓名、身份证号；停保模板至少包含身份证号')

    def cell(row,col):
        return row[col].strip() if col is not None and col<len(row) else ''

    enterprise_cache: dict[str,Enterprise|None]={}
    def resolve_enterprise(raw_name:str) -> tuple[int|None,str|None]:
        if not raw_name: return enterprise_id,None
        if raw_name not in enterprise_cache:
            enterprise_cache[raw_name]=session.scalar(select(Enterprise).where(Enterprise.name==raw_name))
        found=enterprise_cache[raw_name]
        if not found: return None,f'投保单位"{raw_name}"不存在'
        if user.role=='enterprise' and found.id!=user.enterprise_id: return None,'无权为其他投保单位导入数据'
        if found.usage_balance<=0: return None,f'投保单位"{raw_name}"使用费余额不足，请先充值'
        return found.id,None

    errors=[];pending=[];seen=set()
    for row_no,row in enumerate(raw[1:],start=2):
        identity=cell(row,id_col);person_name=cell(row,name_col);phone=cell(row,phone_col)
        row_enterprise_name=cell(row,enterprise_col);row_employer_name=cell(row,employer_col);row_position_name=cell(row,position_col)
        effective_raw=cell(row,effective_col);terminated_raw=cell(row,terminated_col)
        if not identity: errors.append({'row':row_no,'message':'身份证号必填'});continue
        if not is_valid_id_number(identity): errors.append({'row':row_no,'message':'身份证号格式不正确'});continue
        if identity in seen: errors.append({'row':row_no,'message':'表格内身份证号重复'});continue
        seen.add(identity)
        row_enterprise_id,err=resolve_enterprise(row_enterprise_name)
        if err: errors.append({'row':row_no,'message':err});continue
        try:
            effective_at=_parse_business_time(effective_raw,'生效') if effective_raw else None
            terminated_at=_parse_business_time(terminated_raw,'停保') if terminated_raw else None
        except HTTPException as exc:
            errors.append({'row':row_no,'message':exc.detail});continue
        existing=session.scalar(select(InsuredPerson).where(InsuredPerson.enterprise_id==row_enterprise_id,InsuredPerson.id_number==identity))
        if kind=='enrollment':
            if not person_name: errors.append({'row':row_no,'message':'姓名必填'});continue
            if existing and existing.status!='stopped': errors.append({'row':row_no,'message':'该员工已在保或待审核'});continue
            row_position=default_position if row_enterprise_id==enterprise_id and not (row_employer_name or row_position_name) else None
            if row_position is None:
                employer=session.scalar(select(ActualEmployer).where(ActualEmployer.enterprise_id==row_enterprise_id,ActualEmployer.name==row_employer_name)) if row_employer_name else None
                if row_employer_name and not employer: errors.append({'row':row_no,'message':f'实际工作单位"{row_employer_name}"不存在'});continue
                position_query=select(WorkPosition).where(WorkPosition.enterprise_id==row_enterprise_id,WorkPosition.name==row_position_name,WorkPosition.status=='approved')
                if employer: position_query=position_query.where(WorkPosition.actual_employer_id==employer.id)
                row_position=session.scalar(position_query) if row_position_name else None
                if not row_position: errors.append({'row':row_no,'message':'未找到匹配的已审核岗位，请填写实际工作单位与岗位名称，或先在岗位管理中创建并完成审核'});continue
            try:
                if row_position.actual_employer_id is None:
                    if user.role=='enterprise' and not is_enterprise_owner(user): raise HTTPException(403,'岗位未关联实际工作单位，项目负责人无权操作')
                else: assert_employer_access(session,user,row_position.actual_employer_id)
            except HTTPException as exc:
                errors.append({'row':row_no,'message':exc.detail});continue
            if effective_at is not None and terminated_at is not None and terminated_at<=effective_at: errors.append({'row':row_no,'message':'停保日期必须晚于生效日期'});continue
            pending.append(('create',row_enterprise_id,row_position,person_name,identity,phone,effective_at,terminated_at,existing))
        else:
            if not existing: errors.append({'row':row_no,'message':'未找到该单位参保员工'});continue
            if existing.status=='stopped': errors.append({'row':row_no,'message':'该员工已停保'});continue
            try: _person_employer_access(session,user,existing)
            except HTTPException as exc: errors.append({'row':row_no,'message':exc.detail});continue
            pending.append(('stop',row_enterprise_id,None,person_name,identity,phone,None,terminated_at,existing))
    if errors: return {'ok':False,'kind':kind,'success':0,'errors':errors}
    affected=[]
    for action,row_enterprise_id,row_position,person_name,identity,phone,effective_at,terminated_at,existing in pending:
        if action=='create':
            if existing:
                existing.name=person_name;existing.phone=phone;existing.position_id=row_position.id;existing.occupation=row_position.name;existing.occupation_class=row_position.occupation_class;existing.status='pending';item=existing
            else:
                item=InsuredPerson(enterprise_id=row_enterprise_id,position_id=row_position.id,name=person_name,id_number=identity,phone=phone,occupation=row_position.name,occupation_class=row_position.occupation_class,status='pending');session.add(item);session.flush()
            if effective_at is not None:
                member=correct_person_policy_dates(session,item,effective_at,terminated_at)
                if member is not None: item.status='stopped' if member.terminated_at is not None else 'active'
            record_operation(session, user=user, person=item, operation_type='enrollment')
        else:
            item=existing
            if item.status=='active': terminate_person_policy(session,item,terminated_at)
            item.status='stopped'
            record_operation(session, user=user, person=item, operation_type='termination')
        affected.append(item)
    session.commit();audit(session,user,'bulk_enrollment' if kind=='enrollment' else 'bulk_termination','insured_person','',f'count={len(affected)};file={file.filename}')
    return {'ok':True,'kind':kind,'success':len(affected),'errors':[]}
