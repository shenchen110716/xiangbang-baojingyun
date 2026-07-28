"""Aggregate timeliness results into the §13 cards and detail rows.

Reads only — this module computes no verdicts. It extends the pure engine's
twelve counting fields with the four §13.4 metrics that need result rows rather
than status lists: feedback rate, operator-attributable rate, coverage gap and
excess premium.

Scope is applied in the query, behind Phase 1's scope service, so a project
manager's totals are *built from* their authorized employers rather than
filtered afterwards — a filtered-after-the-fact total is how a leak survives a
UI change.
"""
import json
from typing import Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import ActualEmployer, Enterprise, EmploymentTimelinessResult, InsuredPerson, User, WorkPosition
from .employer_scopes import allowed_employer_ids
from .timeliness_engine import summarise

# §13.4 统计卡片至少包括这些。
REQUIRED_CARD_FIELDS = frozenset({
    "enrollment_due", "enrollment_timely", "enrollment_late", "enrollment_missing",
    "termination_due", "termination_timely", "termination_premature",
    "termination_late", "termination_missing",
    "composite_rate", "feedback_rate", "operator_attributable_rate",
    "coverage_gap_seconds", "excess_premium",
})

# 反馈判定为空表示尚未判定（例如无离职事实），不进分母。
_FEEDBACK_DUE = frozenset({"timely", "late", "missing"})
_FEEDBACK_TIMELY = frozenset({"timely"})

# 操作员可归责范围：只有操作员自己能控制的环节才算他的。源头反馈晚、系统处理晚、
# 保司确认晚都不是他的过失，连分母都不该进——否则一个从不误操作的人，会因为别人
# 的延迟而指标下滑（§11.3）。
_OPERATOR_ATTRIBUTABLE_REASONS = frozenset({"normal", "operator_processing_late"})
_TIMELY_STATUSES = frozenset({"timely", "early"})


def _scoped(session: Session, user: User):
    stmt = select(EmploymentTimelinessResult).where(
        EmploymentTimelinessResult.status == "current")
    allowed = allowed_employer_ids(session, user)
    if allowed is not None:
        if not allowed:
            return None
        stmt = stmt.where(EmploymentTimelinessResult.actual_employer_id.in_(allowed))
    if user.role == "enterprise":
        stmt = stmt.where(EmploymentTimelinessResult.enterprise_id == user.enterprise_id)
    return stmt


def _rate(numerator: int, denominator: int) -> Optional[float]:
    """None when nothing is due — an empty project is neither perfect nor
    failing, and 0 or 100 would be read as one of them."""
    if denominator == 0:
        return None
    return round(numerator / denominator * 100, 2)


def scoped_results(session: Session, user: User, **filters) -> list[EmploymentTimelinessResult]:
    stmt = _scoped(session, user)
    if stmt is None:
        return []
    column_filters = {
        "operation_type": EmploymentTimelinessResult.operation_type,
        "timeliness_status": EmploymentTimelinessResult.timeliness_status,
        "responsibility_reason": EmploymentTimelinessResult.responsibility_reason,
        "responsible_user_id": EmploymentTimelinessResult.responsible_user_id,
        "actual_employer_id": EmploymentTimelinessResult.actual_employer_id,
        "enterprise_id": EmploymentTimelinessResult.enterprise_id,
        "feedback_status": EmploymentTimelinessResult.feedback_status,
    }
    for key, column in column_filters.items():
        value = filters.get(key)
        if value is not None and value != "":
            stmt = stmt.where(column == value)

    # 岗位/姓名/身份证号不是结果表自己的字段，要连参保人员表查（用户反馈
    # 2026-07-29：平台端/企业端都要能按这几项查询和统计）。这三项只对已经
    # 匹配到具体人员的记录有意义——待匹配事实本来就不进正式口径，这里查不
    # 到人也在情理之中，不必单独处理。
    position_id = filters.get("position_id")
    person_name = (filters.get("person_name") or "").strip()
    id_number = (filters.get("id_number") or "").strip()
    if position_id or person_name or id_number:
        person_stmt = select(InsuredPerson.id)
        if position_id:
            person_stmt = person_stmt.where(InsuredPerson.position_id == position_id)
        if person_name:
            person_stmt = person_stmt.where(InsuredPerson.name.like(f"%{person_name}%"))
        if id_number:
            person_stmt = person_stmt.where(InsuredPerson.id_number.like(f"%{id_number}%"))
        stmt = stmt.where(EmploymentTimelinessResult.person_id.in_(person_stmt))

    since, until = filters.get("since"), filters.get("until")
    if since:
        stmt = stmt.where(EmploymentTimelinessResult.actual_business_at >= since)
    if until:
        stmt = stmt.where(EmploymentTimelinessResult.actual_business_at <= until)
    return list(session.scalars(stmt.order_by(EmploymentTimelinessResult.id)))


def summary_for(session: Session, user: User, **filters) -> dict:
    rows = scoped_results(session, user, **filters)

    cards = summarise(
        enrollment=[r.timeliness_status for r in rows if r.operation_type == "enrollment"],
        termination=[r.timeliness_status for r in rows if r.operation_type == "termination"],
    )

    feedback_due = [r for r in rows if r.feedback_status in _FEEDBACK_DUE]
    feedback_timely = [r for r in feedback_due if r.feedback_status in _FEEDBACK_TIMELY]

    attributable = [r for r in rows
                    if r.responsibility_reason in _OPERATOR_ATTRIBUTABLE_REASONS]
    attributable_timely = [r for r in attributable
                           if r.timeliness_status in _TIMELY_STATUSES]

    cards.update({
        "feedback_due": len(feedback_due),
        "feedback_timely": len(feedback_timely),
        "feedback_rate": _rate(len(feedback_timely), len(feedback_due)),
        "operator_attributable_due": len(attributable),
        "operator_attributable_rate": _rate(len(attributable_timely), len(attributable)),
        "coverage_gap_seconds": sum(r.coverage_gap_seconds or 0 for r in rows),
        "excess_premium": round(sum(float(r.excess_premium or 0) for r in rows), 2),
        "early_premium": round(sum(float(r.early_premium or 0) for r in rows), 2),
    })
    return cards


def serialize_result(row: EmploymentTimelinessResult) -> dict:
    try:
        evidence = json.loads(row.responsibility_evidence_json or "{}")
    except json.JSONDecodeError:
        evidence = {}
    return {
        "id": row.id,
        "employment_fact_id": row.employment_fact_id,
        "employment_fact_revision_no": row.employment_fact_revision_no,
        "operation_type": row.operation_type,
        "enterprise_id": row.enterprise_id,
        "actual_employer_id": row.actual_employer_id,
        "person_id": row.person_id,
        "responsible_user_id": row.responsible_user_id,
        "actual_business_at": row.actual_business_at,
        "expected_coverage_at": row.expected_coverage_at,
        "actual_coverage_at": row.actual_coverage_at,
        "timeliness_status": row.timeliness_status,
        "delay_seconds": row.delay_seconds,
        "early_seconds": row.early_seconds,
        "coverage_gap_seconds": row.coverage_gap_seconds,
        "excess_premium": row.excess_premium,
        "early_premium": row.early_premium,
        "feedback_status": row.feedback_status,
        "feedback_deadline_at": row.feedback_deadline_at,
        "responsibility_reason": row.responsibility_reason,
        "responsibility_evidence": evidence,
        "product_rule_version": row.product_rule_version,
        "calculation_version": row.calculation_version,
        "calculated_at": row.calculated_at,
    }


def detail_rows(session: Session, user: User, **filters) -> list[dict]:
    """结果表本身只存 id——批量把单位/岗位/人员/责任人的名字补上，明细表格
    才不用只能显示一串数字（用户反馈 2026-07-29：查询/统计要能按这些维度
    看）。批量查而不是逐行查，避免明细行数一多就是 N+1。"""
    from ..core.id_number import mask_id_number

    rows = scoped_results(session, user, **filters)
    person_ids = {r.person_id for r in rows if r.person_id}
    people = {p.id: p for p in session.scalars(
        select(InsuredPerson).where(InsuredPerson.id.in_(person_ids)))} if person_ids else {}
    position_ids = {p.position_id for p in people.values() if p.position_id}
    positions = {p.id: p for p in session.scalars(
        select(WorkPosition).where(WorkPosition.id.in_(position_ids)))} if position_ids else {}
    employer_ids = {r.actual_employer_id for r in rows}
    employers = {e.id: e for e in session.scalars(
        select(ActualEmployer).where(ActualEmployer.id.in_(employer_ids)))} if employer_ids else {}
    enterprise_ids = {r.enterprise_id for r in rows}
    enterprises = {e.id: e for e in session.scalars(
        select(Enterprise).where(Enterprise.id.in_(enterprise_ids)))} if enterprise_ids else {}
    responsible_ids = {r.responsible_user_id for r in rows if r.responsible_user_id}
    responsible_users = {u.id: u for u in session.scalars(
        select(User).where(User.id.in_(responsible_ids)))} if responsible_ids else {}

    payloads = []
    for row in rows:
        person = people.get(row.person_id)
        position = positions.get(person.position_id) if person else None
        employer = employers.get(row.actual_employer_id)
        enterprise = enterprises.get(row.enterprise_id)
        responsible = responsible_users.get(row.responsible_user_id)
        payload = serialize_result(row)
        payload.update({
            "enterprise_name": enterprise.name if enterprise else "",
            "actual_employer_name": employer.name if employer else "",
            "position_name": position.name if position else "",
            "person_name": person.name if person else "",
            "id_number_masked": mask_id_number(person.id_number) if person and person.id_number else "",
            "responsible_user_name": responsible.name if responsible else "",
        })
        payloads.append(payload)
    return payloads


def filter_options(session: Session, user: User) -> dict:
    """按当前用户可见范围，只列出及时率结果里实际出现过的单位/岗位/责任人
    ——下拉选一个查出来空空如也没有意义（用户反馈 2026-07-29）。平台端可
    以看投保单位下拉，企业端已经天然限定在自己单位，不需要这一项。"""
    stmt = _scoped(session, user)
    if stmt is None:
        return {"enterprises": [], "actual_employers": [], "positions": [], "responsible_users": []}
    rows = list(session.scalars(stmt))

    enterprises: list[dict] = []
    if user.role == "admin":
        enterprise_ids = {r.enterprise_id for r in rows}
        if enterprise_ids:
            enterprises = [{"id": e.id, "name": e.name} for e in session.scalars(
                select(Enterprise).where(Enterprise.id.in_(enterprise_ids)).order_by(Enterprise.name))]

    employer_ids = {r.actual_employer_id for r in rows}
    actual_employers = [{"id": e.id, "name": e.name} for e in session.scalars(
        select(ActualEmployer).where(ActualEmployer.id.in_(employer_ids)).order_by(ActualEmployer.name))] if employer_ids else []

    person_ids = {r.person_id for r in rows if r.person_id}
    position_ids: set[int] = set()
    if person_ids:
        position_ids = {pid for pid, in session.execute(
            select(InsuredPerson.position_id).where(
                InsuredPerson.id.in_(person_ids), InsuredPerson.position_id.is_not(None)))}
    positions = [{"id": p.id, "name": p.name} for p in session.scalars(
        select(WorkPosition).where(WorkPosition.id.in_(position_ids)).order_by(WorkPosition.name))] if position_ids else []

    responsible_ids = {r.responsible_user_id for r in rows if r.responsible_user_id}
    responsible_users = [{"id": u.id, "name": u.name} for u in session.scalars(
        select(User).where(User.id.in_(responsible_ids)).order_by(User.name))] if responsible_ids else []

    return {
        "enterprises": enterprises,
        "actual_employers": actual_employers,
        "positions": positions,
        "responsible_users": responsible_users,
    }


# --- §13.4 带审计的 XLSX 导出 -------------------------------------------

_EXPORT_HEADER = [
    "实际工作单位", "姓名", "身份证号", "操作类型", "真实业务时间",
    "应保障时间", "实际保障时间", "及时状态", "延迟秒数", "提前秒数",
    "保障缺口秒数", "额外保费", "反馈状态", "责任原因", "责任人ID",
    "规则版本", "计算版本", "计算时间",
]

_OPERATION_LABEL = {"enrollment": "参保", "termination": "停保"}


def _export_cells(session: Session, row: EmploymentTimelinessResult) -> list:
    from ..core.id_number import id_decrypt, mask_id_number
    from ..models import ActualEmployer, EmploymentFact

    fact = session.get(EmploymentFact, row.employment_fact_id)
    employer = session.get(ActualEmployer, row.actual_employer_id)
    # 身份证原文不得出现在导出中（§15）；只从密文解出后立即脱敏。
    masked = ""
    if fact and fact.id_number_cipher:
        masked = mask_id_number(id_decrypt(fact.id_number_cipher))

    def iso(value):
        return value.isoformat(sep=" ")[:19] if value else ""

    return [
        employer.name if employer else "",
        fact.person_name if fact else "",
        masked,
        _OPERATION_LABEL.get(row.operation_type, row.operation_type),
        iso(row.actual_business_at),
        iso(row.expected_coverage_at),
        iso(row.actual_coverage_at),
        row.timeliness_status,
        row.delay_seconds,
        row.early_seconds,
        row.coverage_gap_seconds,
        float(row.excess_premium or 0),
        row.feedback_status,
        row.responsibility_reason,
        row.responsible_user_id or "",
        row.product_rule_version,
        row.calculation_version,
        iso(row.calculated_at),
    ]


def build_export(session: Session, user: User, *, filters: dict) -> tuple[bytes, dict]:
    """Returns (xlsx_bytes, audit_metadata).

    The digest is taken over the bytes actually returned, so the audit row
    proves what was handed out rather than what we intended to hand out
    (§13.4).
    """
    import hashlib
    import io as _io

    import openpyxl

    from ..core.business_time import business_now

    rows = scoped_results(session, user, **filters)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "及时率明细"
    sheet.append(_EXPORT_HEADER)
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DCE6FF")
    for row in rows:
        sheet.append(_export_cells(session, row))
    for index in range(2, sheet.max_row + 1):
        # Text format, or Excel renders a masked ID as a number/date.
        sheet.cell(index, 3).number_format = "@"
    for column, width in {"A": 22, "B": 12, "C": 22, "D": 10, "E": 20, "F": 20,
                          "G": 20, "H": 12, "N": 24, "R": 20}.items():
        sheet.column_dimensions[column].width = width

    output = _io.BytesIO()
    book.save(output)
    book.close()
    payload = output.getvalue()

    meta = {
        "filters": {k: v for k, v in (filters or {}).items() if v not in (None, "")},
        "exported_by": user.id,
        "exported_at": business_now().isoformat(),
        "row_count": len(rows),
        "file_digest": hashlib.sha256(payload).hexdigest(),
    }
    return payload, meta
