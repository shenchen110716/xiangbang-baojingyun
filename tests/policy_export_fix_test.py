"""Regression: 保单人员明细导出/打印证明为空.

Root cause: export_policy() (routers/reports.py) and the miniprogram/web
certificate flow both used InsuredPerson.policy_id to find "who is on this
policy". That column gets cleared or repointed on stop/renew — PolicyMember
is the authoritative source (same invariant services/policies.py:policy_dict
already relies on, per CLAUDE.md's "停保以当前有效 PolicyMember 保障期为权威").
This test builds exactly the failure shape: a person whose PolicyMember row
still points at the policy but whose InsuredPerson.policy_id has been
cleared, and proves both the export endpoint and the /insured?policy_id=
filter (used by CertificateView.vue for 打印证明) still find them.
"""
import os
import sys
import tempfile
from datetime import timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


def run():
    with tempfile.TemporaryDirectory(prefix="xbb-policy-export-fix-") as folder:
        os.environ["DATABASE_URL"] = f"sqlite:///{Path(folder) / 'test.db'}"
        os.environ["ADMIN_PASSWORD"] = "admin123"
        os.environ["ENTERPRISE_PASSWORD"] = "enterprise123"

        from openpyxl import load_workbook
        from sqlalchemy import select

        from backend.app import startup
        from backend.core.business_time import business_now
        from backend.core.db import SessionLocal
        from backend.models import (
            ActualEmployer, Enterprise, InsurancePlan, InsuredPerson, Policy,
            PolicyMember, User, WorkPosition,
        )
        from backend.routers.insured import insured as list_insured
        from backend.routers.reports import export_policy

        startup()
        with SessionLocal() as session:
            admin = session.scalar(select(User).where(User.role == "admin"))

            enterprise = Enterprise(name="导出修复测试企业", kind="企业", contact="", phone="", status="active")
            session.add(enterprise); session.flush()
            employer = ActualEmployer(enterprise_id=enterprise.id, name="测试用工单位")
            session.add(employer); session.flush()
            plan = InsurancePlan(insurer="测试保司", name="测试方案", price=100, commission_rate=.2,
                                  effective_mode="immediate", billing_mode="daily", status="active")
            session.add(plan); session.flush()
            position = WorkPosition(enterprise_id=enterprise.id, actual_employer_id=employer.id,
                                     name="测试岗位", occupation_class="1-4类", plan_id=plan.id, status="approved")
            session.add(position); session.flush()
            policy = Policy(policy_no="POL-EXPORT-FIX-TEST", enterprise_id=enterprise.id, plan_id=plan.id, status="active")
            session.add(policy); session.flush()

            # 关键：person.policy_id 从一开始就不指向这份保单（None），只有
            # PolicyMember 挂着——模拟停保/续保后 InsuredPerson.policy_id 被
            # 清空/改指别处，但这份保单当时的在保记录（PolicyMember）还在的真实场景。
            person = InsuredPerson(enterprise_id=enterprise.id, name="导出修复测试员工",
                                    id_number="110101199001010011", position_id=position.id,
                                    occupation_class="1-4类", status="active", policy_id=None)
            session.add(person); session.flush()
            session.add(PolicyMember(policy_id=policy.id, person_id=person.id,
                                      effective_at=business_now() - timedelta(days=1), status="active"))
            session.commit(); session.refresh(person)

            assert person.policy_id is None, "测试前提：InsuredPerson.policy_id 必须是空的"

            # 1. /policies/{id}/export 之前会因为按 InsuredPerson.policy_id 查询而导出空表。
            import asyncio
            import io

            response = export_policy(policy.id, admin, session)

            async def _collect():
                out = b""
                async for c in response.body_iterator:
                    out += c if isinstance(c, bytes) else c.encode()
                return out

            content = asyncio.run(_collect())
            book = load_workbook(filename=io.BytesIO(content))
            sheet = book.active
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            assert len(rows) == 1, f"导出应该有 1 行人员明细，实际 {len(rows)} 行：{rows}"
            assert rows[0][5] == "导出修复测试员工", f"导出的姓名列不对：{rows[0]}"

            # 2. GET /insured?policy_id=X 之前没有这个参数，前端只能自己按
            #    policy_id 过滤 /insured 全量结果，同样的空字段问题会导致
            #    "打印证明"也是空的。
            scoped = list_insured("", policy.id, admin, session)
            assert len(scoped) == 1, f"/insured?policy_id= 应该返回 1 人，实际 {len(scoped)} 人"
            assert scoped[0]["name"] == "导出修复测试员工"

            # 3. 不传 policy_id 或传一个不相关的保单 id，不应该混进这个人。
            unrelated_policy = Policy(policy_no="POL-UNRELATED", enterprise_id=enterprise.id, plan_id=plan.id, status="active")
            session.add(unrelated_policy); session.commit()
            unrelated_scoped = list_insured("", unrelated_policy.id, admin, session)
            assert unrelated_scoped == [], f"不相关保单的 policy_id 过滤不应该查到人，实际 {unrelated_scoped}"

    print("policy export fix test: PASS")


if __name__ == "__main__":
    run()
