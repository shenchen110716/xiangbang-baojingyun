"""Reporting aggregation and audited export (v4.2 §13, §15).

Phase 3's engine returns the twelve counting fields; §13.4 requires four more
that need result rows rather than status lists — feedback rate, operator-
attributable rate, coverage gap and excess premium. This is where they live.

Two rules are load-bearing here:
- 身份证原文不得出现在页面、导出或行级错误文件中 (§15).
- 跨企业导出必须记录筛选条件、导出人、时间、行数和文件摘要 (§13.4).
"""
import io
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

os.environ.setdefault("ID_ENCRYPTION_KEY", "reporting-test-key")

from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.core.db import Base
from backend.core.id_number import id_encrypt, id_hash
from backend.models import (
    ActualEmployer,
    AuditLog,
    EmploymentFact,
    EmploymentTimelinessResult,
    Enterprise,
    InsuredPerson,
    User,
    UserEmployerScope,
    WorkPosition,
)
from backend.services.timeliness_reporting import (
    REQUIRED_CARD_FIELDS,
    build_export,
    filter_options,
    summary_for,
    detail_rows,
)

RAW_ID = "340123199001011238"
MASKED_ID = "340123********1238"


def _now():
    return datetime.now(timezone.utc)


class _Ctx:
    pass


def _setup(session) -> _Ctx:
    ctx = _Ctx()
    ctx.enterprise = Enterprise(name="报表企业")
    session.add(ctx.enterprise)
    session.flush()
    ctx.employer_a = ActualEmployer(enterprise_id=ctx.enterprise.id, name="项目 A")
    ctx.employer_b = ActualEmployer(enterprise_id=ctx.enterprise.id, name="项目 B")
    session.add_all([ctx.employer_a, ctx.employer_b])
    session.flush()
    ctx.owner = User(username="rep_owner", password_hash="x", name="主管",
                     role="enterprise", enterprise_id=ctx.enterprise.id,
                     enterprise_role="owner", is_owner=True)
    ctx.manager = User(username="rep_manager", password_hash="x", name="负责人",
                       role="enterprise", enterprise_id=ctx.enterprise.id,
                       enterprise_role="project_manager", is_owner=False)
    ctx.admin = User(username="rep_admin", password_hash="x", name="平台管理员", role="admin")
    session.add_all([ctx.owner, ctx.manager, ctx.admin])
    session.flush()
    session.add(UserEmployerScope(
        user_id=ctx.manager.id, enterprise_id=ctx.enterprise.id,
        actual_employer_id=ctx.employer_a.id, responsibility_type="primary",
        granted_by=ctx.owner.id, status="active", assigned_at=_now()))
    session.flush()
    return ctx


def _fact(session, ctx, employer):
    f = EmploymentFact(
        enterprise_id=ctx.enterprise.id, actual_employer_id=employer.id,
        person_name="张三", id_number_hash=id_hash(RAW_ID),
        id_number_cipher=id_encrypt(RAW_ID),
        actual_hire_at=datetime(2026, 3, 1, tzinfo=timezone.utc),
        status="active", revision_no=1, created_at=_now())
    session.add(f)
    session.flush()
    return f


def _person(session, ctx, *, employer, name, id_number):
    position = WorkPosition(enterprise_id=ctx.enterprise.id, actual_employer_id=employer.id,
                            name=f"{name}的岗位", occupation_class="1-3类")
    session.add(position); session.flush()
    person = InsuredPerson(enterprise_id=ctx.enterprise.id, name=name,
                           id_number=id_number, position_id=position.id, status="active")
    session.add(person); session.flush()
    return person


def _result(session, ctx, *, employer=None, operation_type="enrollment",
            status="timely", feedback="timely", reason="normal",
            gap=0, excess=0.0, responsible=None, person=None):
    employer = employer or ctx.employer_a
    fact = _fact(session, ctx, employer)
    row = EmploymentTimelinessResult(
        employment_fact_id=fact.id, employment_fact_revision_no=1,
        operation_type=operation_type, enterprise_id=ctx.enterprise.id,
        actual_employer_id=employer.id,
        person_id=person.id if person else None,
        responsible_user_id=responsible,
        timeliness_status=status, feedback_status=feedback,
        responsibility_reason=reason, coverage_gap_seconds=gap,
        excess_premium=excess, product_rule_version=1, calculation_version=1,
        calculated_at=_now(), status="current")
    session.add(row)
    session.flush()
    return row


def run() -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    _test_stat_cards_cover_every_required_metric(engine)
    _test_feedback_rate_counts_only_judged_feedback(engine)
    _test_operator_attributable_rate_excludes_other_parties(engine)
    _test_gap_and_excess_are_summed(engine)
    _test_manager_summary_counts_only_authorized_employers(engine)
    _test_filters_narrow_details(engine)
    _test_person_level_filters_narrow_details(engine)
    _test_enterprise_filter_only_for_admin(engine)
    _test_filter_options_lists_only_whats_in_scope(engine)
    _test_detail_rows_include_display_names(engine)
    _test_export_masks_identity(engine)
    _test_export_digest_and_row_count_match_the_file(engine)

    print("timeliness reporting tests passed")


def _test_stat_cards_cover_every_required_metric(engine):
    """§13.4 统计卡片至少包括这些。"""
    with Session(engine) as session:
        ctx = _setup(session)
        _result(session, ctx)
        s = summary_for(session, ctx.owner)
        assert set(s) >= REQUIRED_CARD_FIELDS, sorted(REQUIRED_CARD_FIELDS - set(s))
    print("  stat cards complete ok")


def _test_feedback_rate_counts_only_judged_feedback(engine):
    with Session(engine) as session:
        ctx = _setup(session)
        _result(session, ctx, feedback="timely")
        _result(session, ctx, feedback="late")
        _result(session, ctx, feedback="")        # 未判定：不进分母
        s = summary_for(session, ctx.owner)
        assert s["feedback_rate"] == 50.0, s["feedback_rate"]
    print("  feedback rate ok")


def _test_operator_attributable_rate_excludes_other_parties(engine):
    """操作员可归责及时率：只把操作员自己能控制的延迟算作其过失。"""
    with Session(engine) as session:
        ctx = _setup(session)
        _result(session, ctx, status="timely", reason="normal")
        _result(session, ctx, status="late", reason="operator_processing_late")
        # 源头反馈晚、保司确认晚：不是操作员的锅，不计入其可归责分子的失败
        _result(session, ctx, status="late", reason="source_feedback_late")
        _result(session, ctx, status="late", reason="insurer_confirmation_late")
        s = summary_for(session, ctx.owner)
        # 4 个事件里，操作员可归责范围 = timely(1) + operator_late(1) = 2；其中及时 1
        assert s["operator_attributable_rate"] == 50.0, s["operator_attributable_rate"]
    print("  operator attributable rate ok")


def _test_gap_and_excess_are_summed(engine):
    with Session(engine) as session:
        ctx = _setup(session)
        _result(session, ctx, gap=3 * 86400, excess=12.5)
        _result(session, ctx, gap=86400, excess=7.5)
        s = summary_for(session, ctx.owner)
        assert s["coverage_gap_seconds"] == 4 * 86400
        assert s["excess_premium"] == 20.0
    print("  gap and excess ok")


def _test_manager_summary_counts_only_authorized_employers(engine):
    with Session(engine) as session:
        ctx = _setup(session)
        _result(session, ctx, employer=ctx.employer_a)
        _result(session, ctx, employer=ctx.employer_b)
        assert summary_for(session, ctx.owner)["enrollment_due"] == 2
        assert summary_for(session, ctx.manager)["enrollment_due"] == 1, \
            "项目负责人的总数必须只由其授权单位构成"
    print("  manager scope ok")


def _test_filters_narrow_details(engine):
    """§13.1 筛选：操作员、单位、参保/停保、及时状态、责任原因。"""
    with Session(engine) as session:
        ctx = _setup(session)
        _result(session, ctx, status="late", reason="operator_processing_late",
                responsible=ctx.owner.id)
        _result(session, ctx, status="timely")
        _result(session, ctx, operation_type="termination", status="late")

        rows = detail_rows(session, ctx.owner, operation_type="enrollment",
                           timeliness_status="late",
                           responsibility_reason="operator_processing_late")
        assert len(rows) == 1, rows
        assert all(r["operation_type"] == "enrollment"
                   and r["timeliness_status"] == "late" for r in rows)

        assert len(detail_rows(session, ctx.owner, responsible_user_id=ctx.owner.id)) == 1
        assert len(detail_rows(session, ctx.owner,
                               actual_employer_id=ctx.employer_b.id)) == 0
    print("  filters ok")


def _test_person_level_filters_narrow_details(engine):
    """岗位/姓名/身份证号查询（用户反馈 2026-07-29：平台端/企业端都要能按
    这几项查询和统计）。"""
    with Session(engine) as session:
        ctx = _setup(session)
        zhang = _person(session, ctx, employer=ctx.employer_a, name="张三", id_number="340123199001011238")
        li = _person(session, ctx, employer=ctx.employer_b, name="李四", id_number="110101199003077715")
        _result(session, ctx, employer=ctx.employer_a, person=zhang)
        _result(session, ctx, employer=ctx.employer_b, person=li)

        assert len(detail_rows(session, ctx.owner, position_id=zhang.position_id)) == 1
        assert len(detail_rows(session, ctx.owner, person_name="张")) == 1
        assert len(detail_rows(session, ctx.owner, person_name="不存在的名字")) == 0
        assert len(detail_rows(session, ctx.owner, id_number="340123")) == 1
        assert len(detail_rows(session, ctx.owner, id_number="110101")) == 1
    print("  person-level filters ok")


def _test_enterprise_filter_only_for_admin(engine):
    """投保单位查询是平台端的能力；企业端本来就锁死在自己单位，传别的
    enterprise_id 也不能看到别家（不是新增泄露面，只是确认没打开）。"""
    with Session(engine) as session:
        ctx = _setup(session)
        other = Enterprise(name="报表企业-乙")
        session.add(other); session.flush()
        other_employer = ActualEmployer(enterprise_id=other.id, name="乙项目")
        session.add(other_employer); session.flush()

        _result(session, ctx, employer=ctx.employer_a)
        fact = EmploymentFact(
            enterprise_id=other.id, actual_employer_id=other_employer.id,
            person_name="王五", id_number_hash=id_hash("110101199003077715"),
            id_number_cipher=id_encrypt("110101199003077715"),
            actual_hire_at=datetime(2026, 3, 1, tzinfo=timezone.utc),
            status="active", revision_no=1, created_at=_now())
        session.add(fact); session.flush()
        session.add(EmploymentTimelinessResult(
            employment_fact_id=fact.id, employment_fact_revision_no=1,
            operation_type="enrollment", enterprise_id=other.id,
            actual_employer_id=other_employer.id, timeliness_status="timely",
            feedback_status="timely", responsibility_reason="normal",
            product_rule_version=1, calculation_version=1, calculated_at=_now(),
            status="current"))
        session.flush()

        assert len(detail_rows(session, ctx.admin, enterprise_id=ctx.enterprise.id)) == 1
        assert len(detail_rows(session, ctx.admin, enterprise_id=other.id)) == 1
        assert len(detail_rows(session, ctx.admin)) == 2, "不传 enterprise_id 时管理员应该看到全部"
        assert len(detail_rows(session, ctx.owner, enterprise_id=other.id)) == 0, \
            "企业端传别家 enterprise_id 不应该看到别家数据"
    print("  enterprise filter admin-only scoping ok")


def _test_filter_options_lists_only_whats_in_scope(engine):
    with Session(engine) as session:
        ctx = _setup(session)
        zhang = _person(session, ctx, employer=ctx.employer_a, name="张三", id_number="340123199001011238")
        _result(session, ctx, employer=ctx.employer_a, person=zhang, responsible=ctx.owner.id)

        admin_options = filter_options(session, ctx.admin)
        assert any(e["id"] == ctx.enterprise.id for e in admin_options["enterprises"]), \
            "管理员应该能看到这家投保单位"
        assert any(e["id"] == ctx.employer_a.id for e in admin_options["actual_employers"])
        assert not any(e["id"] == ctx.employer_b.id for e in admin_options["actual_employers"]), \
            "没有结果的工作单位不该出现在下拉里"
        assert any(p["id"] == zhang.position_id for p in admin_options["positions"])
        assert any(u["id"] == ctx.owner.id for u in admin_options["responsible_users"])

        owner_options = filter_options(session, ctx.owner)
        assert owner_options["enterprises"] == [], "企业端不需要投保单位下拉"
    print("  filter options scoping ok")


def _test_detail_rows_include_display_names(engine):
    """明细行要能直接显示姓名/岗位/单位，不是一串裸 id（用户反馈
    2026-07-29）。"""
    with Session(engine) as session:
        ctx = _setup(session)
        zhang = _person(session, ctx, employer=ctx.employer_a, name="张三", id_number="340123199001011238")
        _result(session, ctx, employer=ctx.employer_a, person=zhang, responsible=ctx.owner.id)

        rows = detail_rows(session, ctx.owner)
        assert len(rows) == 1
        row = rows[0]
        assert row["enterprise_name"] == ctx.enterprise.name
        assert row["actual_employer_name"] == ctx.employer_a.name
        assert row["person_name"] == "张三"
        assert row["position_name"] == "张三的岗位"
        assert row["id_number_masked"] == "340123********1238"
        assert row["responsible_user_name"] == ctx.owner.name
    print("  detail rows carry display names ok")


def _test_export_masks_identity(engine):
    """§15 身份证原文不得出现在导出中。"""
    from openpyxl import load_workbook
    with Session(engine) as session:
        ctx = _setup(session)
        _result(session, ctx)
        payload, meta = build_export(session, ctx.owner, filters={})
        book = load_workbook(io.BytesIO(payload))
        cells = [str(c.value) for row in book.active.iter_rows() for c in row]
        joined = "\n".join(cells)
        assert RAW_ID not in joined, "导出不得含身份证原文"
        assert MASKED_ID in joined, "导出应含脱敏身份证"
    print("  export masks identity ok")


def _test_export_digest_and_row_count_match_the_file(engine):
    """§13.4 导出必须记录筛选条件、导出人、时间、行数和文件摘要。"""
    import hashlib
    with Session(engine) as session:
        ctx = _setup(session)
        _result(session, ctx)
        _result(session, ctx)
        payload, meta = build_export(session, ctx.owner,
                                     filters={"operation_type": "enrollment"})
        assert meta["row_count"] == 2, meta
        assert meta["file_digest"] == hashlib.sha256(payload).hexdigest(), \
            "摘要必须是真实文件内容的摘要"
        assert meta["filters"]["operation_type"] == "enrollment"
        assert meta["exported_by"] == ctx.owner.id
        assert meta["exported_at"]
    print("  export audit metadata ok")


if __name__ == "__main__":
    run()
